import paramiko
import re
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor

# Function to connect to a device and retrieve associated IP addresses
def get_associated_ips(ip_address, username, password):
    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(ip_address, username=username, password=password)

        stdin, stdout, stderr = client.exec_command('show ip interface brief | exclude una')
        output = stdout.read().decode()

        ip_pattern = re.compile(r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b')
        ip_addresses = ip_pattern.findall(output)

        client.close()
        return ip_addresses
    except paramiko.AuthenticationException:
        print(f"Authentication failed for {ip_address}")
        return ['Authentication Failed']
        
    except paramiko.SSHException:
        print(f"Failed to establish SSH connection to {ip_address}")
        return ['SSH Connection Failed']
        
    except paramiko.ssh_exception.NoValidConnectionsError:
        print(f"Unable to connect to {ip_address} on port 22")
        return ['Failed to connect to port 22']

    except Exception as e:
        print(f"An error occurred while retrieving IPs from {ip_address}: {e}")
        return ['Unknown error occurred']
            
    

# Function to process the Excel sheet and update associated IPs for a specific row range
def update_associated_ips(sheet, username, password, min_row, max_row):
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=False):
        name = row[0].value
        sys_class_name = row[1].value
        model_id = row[2].value
        ip_address = row[3].value
        associated_ip_cell = row[4]  # The cell object for 'associated_ip' column

        if ip_address:
            print(f"Fetching associated IPs for {name} (IP: {ip_address})")
            associated_ips = get_associated_ips(ip_address, username, password)
            associated_ip_cell.value = ', '.join(associated_ips)
        else:
            print(f"No IP address found for {name}")

# Function to run multithreading for different row ranges
def process_in_parallel(file_path, username, password, row_ranges):
    # Load the workbook once
    wb = load_workbook(file_path)
    sheet = wb['non_cisco_new']

    # Use a ThreadPoolExecutor for parallel execution
    with ThreadPoolExecutor() as executor:
        futures = [
            executor.submit(update_associated_ips, sheet, username, password, min_row, max_row)
            for min_row, max_row in row_ranges
        ]
        for future in futures:
            future.result()  # Wait for all threads to complete

    # Save the workbook after all threads are done
    wb.save(file_path)
    print(f"Workbook updated successfully for all ranges.")



# Input details
file_path = 'Cisco and Aruba switches_pending.xlsx'
username = 'pankajp'
password = 'PanSum4444'

# Define row ranges for multithreading
row_ranges = [(1, 500),(500, 1000),(1000, 1500),(1500, 2000), (2000, 2500),(2500, 3000), (3000, 3500), (3500, 3832)]  # You can add more ranges
#row_ranges = [(2000, 2025)]
# Call the function to process rows in parallel
process_in_parallel(file_path, username, password, row_ranges)
